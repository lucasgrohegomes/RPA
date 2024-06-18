import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import selenium
from selenium.webdriver import Chrome
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
import pyautogui as p
import rpa as r
import pygetwindow as pw
import datetime as dt
import openpyxl.drawing.image

# wb = Workbook()
# nome_arquivo = 'meu_arquivo_teste.xlsx'
# ws1 = wb.active
# ws1.title = 'planilha1'
#
# wb.save(filename= r'C:\Users\lucas.gomes\Desktop\meu_arquivo_teste.xlsx')

# r.init(visual_automation=True, chrome_browser=True)
#
# r.url(r'https://br.investing.com/currencies/usd-brl')
# janela = pw.getActiveWindow()
# janela.maximize()
# p.sleep(3)
# dolar = r.read('//*[@id="__next"]/div[2]/div[2]/div[2]/div[1]/div[1]/div[3]/div[1]/div[1]/div[1]')
# p.sleep(1)
# # print('A cotação do dolar é: {}, data e hora: {}'.format(dolar, str(dt.datetime.now())))
#
# r.close()
#
# ws1['A7'] = 'A cotação do dolar hoje, %s, é: ' % dt.date.today().strftime('%d/%m/%Y')
# ws1['B7'] = dolar
# ws1['B7'].fill = PatternFill('solid', fgColor = '00FFFF00')
#
# wb.save(filename= r'C:\Users\lucas.gomes\Desktop\meu_arquivo_teste.xlsx')

# navegador = Chrome()
# navegador.get(r'https://br.investing.com/currencies/usd-brl')
# p.sleep(1)
# navegador.maximize_window()
# p.sleep(1)
# dolar = navegador.find_element('xpath', '//*[@id="__next"]/div[2]/div[2]/div[2]/div[1]/div[1]/div[3]/div[1]/div[1]/div[1]').text
# p.sleep(1)
# navegador.close()
#
# ws1['A9'] = 'A cotação do dolar hoje, %s, é: ' % dt.date.today().strftime('%d/%m/%Y')
# ws1['B9'] = dolar
# ws1['B9'].fill = PatternFill('solid', fgColor='00FFFF00')
# wb.save(filename = r'C:\Users\lucas.gomes\Desktop\meu_arquivo_teste.xlsx')

wb = load_workbook(filename=r'C:\Users\lucas.gomes\Desktop\meu_arquivo_teste.xlsx')
ws1 = wb['planilha1']

wb.save(filename=r'C:\Users\lucas.gomes\Desktop\meu_arquivo_teste.xlsx')

navegador = Chrome()
navegador.get('https://crescemosjuntos.com.br/trabalhe-conosco/vaga?protocolo=00240/2023')
navegador.maximize_window()

p.sleep(3)
with open('logo.png', 'wb') as file:
    logo = navegador.find_element('xpath', '/html/body/div[1]/header/div[2]/div[3]/div/div/div/div/a/div/img')
    file.write(logo.screenshot_as_png)

navegador.close()

img = openpyxl.drawing.image.Image('logo.png')
ws1.add_image(img, 'A1')
wb.save(filename=r'C:\Users\lucas.gomes\Desktop\meu_arquivo_teste.xlsx')